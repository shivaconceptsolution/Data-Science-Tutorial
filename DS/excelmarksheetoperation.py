#!/usr/bin/env python
# coding: utf-8

# In[12]:


import openpyxl
import numpy as np
data=openpyxl.load_workbook("P:\\excel\\marksheetopall.xlsx")
sheet=data.active
max_column=sheet.max_column
max_row=sheet.max_row

l=[]
n1=[]
n2=[]
n3=[]
n4=[]
n5=[]
#display record of max and min obtained mark student

for i in range(1,max_row+1):
    for j in range(1,max_column+1):
        cell=sheet.cell (row=i,column=j)
        l.append(cell.value)
print("que(1),max and min marks student")
l1=[]
for i in range(2,(max_row)+1):
    cell=sheet.cell(row=i,column=13)
    l1.append(cell.value)
print("max mark:-",max(l1))
print("min mark:-",min(l1))
print("                      roll no. name   branch  sem1 sem2 sem3 sem4 sem5 sem6 sem7 sem8 gender total per%")

for i in range(0,len(l),14):
    a1=l[0+i:14+i]
    if 3999 in a1:
        print("max student details:- ",a1)
    if 3225 in a1:
        print("min student details:- ",a1)

print("que(2),all student percentage:-")
for i in range(0,len(l),14):
    a1=l[0+i:14+i]
    
    a5=a1[1]
    a6=a1[13]
    
    print(a5,"  ",a6)
print("girls percentage all branch:-")
print("name   branch   gender total   per%")
for i in range(0,len(l),14):
    a1=l[0+i:14+i]
    if 'f' in a1:
        a9=a1[1]
        a2=a1[2]
        a3=a1[11]
        a4=a1[12]
        a5=a1[13]

        print(a9,"  ",a2,"   ",a3,"  ",a4,"  ",a5)
print("boys percentage all branch:-")
print("name branch   gender total   per%")
for i in range(0,len(l),14):
    a1=l[0+i:14+i]
    if 'm' in a1:
        a9=a1[1]
        a2=a1[2]
        a3=a1[11]
        a4=a1[12]
        a5=a1[13]

        print(a9,"  ",a2,"  ",a3,"  ",a4,"  ",a5)
print("que(3),branch wise percentage:-")
print("mechanical branch")
print("name      branch   gender  total    per%")
for i in range(0,len(l),14):
    a1=l[0+i:14+i]

    if 'mech.' in a1:
        
        a9=a1[1]
        a2=a1[2]
        a3=a1[11]
        a4=a1[12]
        a5=a1[13]
        n1.append(a4)
        print(a9,"  ",a2,"  ",a3,"  ",a4,"  ",a5)
        m1=max(n1)
print("chemical  branch")
print("name    branch gender  total    per%")
for i in range(0,len(l),14):
    a1=l[0+i:14+i]
  
    if 'chem.' in a1:
      
        a9=a1[1]
        a2=a1[2]
        a3=a1[11]
        a4=a1[12]
        a5=a1[13]
        n2.append(a4)
        print(a9,"  ",a2,"  ",a3,"  ",a4,"  ",a5)
        m2=max(n2)
print("cs  branch")
print("name  branch   gender  total    per%")
for i in range(0,len(l),14):
    a1=l[0+i:14+i]
  
    if 'cs' in a1:
      
        a9=a1[1]
        a2=a1[2]
        a3=a1[11]
        a4=a1[12]
        a5=a1[13]
        n3.append(a4)
        print(a9,"  ",a2,"  ",a3,"  ",a4,"  ",a5)
        m3=max(n3)
print("it  branch")
print("name branch  gender  total  per%")
for i in range(0,len(l),14):
    a1=l[0+i:14+i]
   
    if 'it' in a1:
      
        a9=a1[1]
        a2=a1[2]
        a3=a1[11]
        a4=a1[12]
        a5=a1[13]
        n4.append(a4)
        print(a9,"  ",a2,"  ",a3,"  ",a4,"  ",a5)
        m4=max(n4)
print("civil  branch")
print("name  branch   gender  total  per%")
for i in range(0,len(l),14):
    a1=l[0+i:14+i]

    if 'civil' in a1:
      
        a9=a1[1]
        a2=a1[2]
        a3=a1[11]
        a4=a1[12]
        a5=a1[13]
        n5.append(a4)
        print(a9,"  ",a2,"  ",a3,"  ",a4,"  ",a5)
        m5=max(n5)
print(" max marks for each branch and semester for male , female and all")
print("mechanical max percent:-")
print("roll no. name   branch  sem1 sem2 sem3 sem4 sem5 sem6 sem7 sem8 gender total per%")
for i in range(0,len(l),14):
    a1=l[0+i:14+i]
    if m1 in a1:
        print(a1)
print("chemical max percent:-")
print("roll no. name   branch  sem1 sem2 sem3 sem4 sem5 sem6 sem7 sem8 gender total per%")
for i in range(0,len(l),14):
    a1=l[0+i:14+i]
    if m2 in a1:
        print(a1)
print("cs max percent:-")
print("roll no. name   branch  sem1 sem2 sem3 sem4 sem5 sem6 sem7 sem8 gender total per%")
for i in range(0,len(l),14):
    a1=l[0+i:14+i]
    if m3 in a1:
        print(a1)
print("it max percent:-")
print("roll no. name   branch  sem1 sem2 sem3 sem4 sem5 sem6 sem7 sem8 gender total per%")
for i in range(0,len(l),14):
    a1=l[0+i:14+i]
    if m4 in a1:
        print(a1)
print("civil max percent:-")
print("roll no. name   branch  sem1 sem2 sem3 sem4 sem5 sem6 sem7 sem8 gender total per%")
for i in range(0,len(l),14):
    a1=l[0+i:14+i]
    if m5 in a1:
        print(a1)


